from pathlib import Path
from collections import namedtuple
from openpyxl import Workbook
from django.test import TestCase

from ..models import Client, Place, Contract, Manufacturer
from ..loader import CILoader
from ..cis_mapping import CLIENT_CELL, SUMMARY_SHEET, CIS_SHEET, \
    APPLIANCES_SHEET

SPREADSHEET_FILE = 'cis_test.xlsx'
CLIENT_NAME = 'New Client'


class CILoaderTest(TestCase):

    @classmethod
    def setUpTestData(cls):
        create_workbook()
        cls.loader = CILoader(SPREADSHEET_FILE).save()

    @classmethod
    def tearDownClass(cls):
        Path(SPREADSHEET_FILE).unlink()
        super().tearDownClass()

    def test_return_correct_client_object(self):
        self.assertIsInstance(self.loader.client, Client)
        self.assertEqual(self.loader.client.name, CLIENT_NAME)

    def test_return_correct_site_objects(self):
        sites = self.loader.sites
        keys = {'NY1', 'NY2', 'SP', 'BH'}
        for k, site in sites.items():
            self.assertIn(k, keys)
            self.assertIsInstance(site, Place)

    def test_return_correct_contract_objects(self):
        contracts = self.loader.contracts
        keys = {'SP-001', 'BH-001', 'NY-001', 'NY-002'}
        for k, contract in contracts.items():
            self.assertIn(k, keys)
            self.assertIsInstance(contract, Contract)

    def test_return_correct_manufacture_objects(self):
        manufacturers = self.loader.manufacturers
        keys = {'Cisco', 'F5'}
        for k, manufacturer in manufacturers.items():
            self.assertIn(k, keys)
            self.assertIsInstance(manufacturer, Manufacturer)

    def test_loader_contains_correct_number_of_cis(self):
        self.assertEqual(len(self.loader.cis), 5)

    def test_errors_contain_duplicated_items(self):
        create_workbook()
        loader = CILoader(SPREADSHEET_FILE).save()
        self.assertEqual(len(loader.errors), 5)
        self.assertEqual(len(loader.cis), 0)
        self.assertTrue(
            'unique constraint' in str(loader.errors[0].exc).lower()
        )


def create_workbook():
    wb = Workbook()
    set_summary_sheet(wb)
    set_cis_sheet(wb)
    set_appliances_sheet(wb)
    wb.save(filename=SPREADSHEET_FILE)


def set_summary_sheet(workbook):
    summary_sheet = workbook.active
    summary_sheet.title = SUMMARY_SHEET
    summary_sheet[CLIENT_CELL] = CLIENT_NAME


def set_cis_sheet(workbook):
    cis_sheet = workbook.create_sheet(CIS_SHEET)
    Row = namedtuple('Row', (
        'hostname',
        'ip',
        'description',
        'deployed',
        'business_impact',
        'site',
        'site_description',
        'contract',
        'contract_begin',
        'contract_end',
        'contract_description',
        'username',
        'password',
        'enable_password',
        'instructions',
    ))
    base = Row('router_sp', '172.16.5.10', 'Main Router',
               'x', 'high', 'SP', 'Center', 'SP-001',
               '2021-01-01', '2022-01-01', 'Contract Details',
               'admin', 'admin', 'enable', 'Instructions')
    rows = (
        Row._fields,
        base,
        base._replace(hostname='router_bh', ip='172.16.6.10',
                      site='BH', contract='BH-001'),
        base._replace(hostname='wlc1', ip='172.16.10.10',
                      description='Controller Floor 1', site='NY1',
                      site_description='Main', contract='NY-001'),
        base._replace(hostname='wlc2', ip='172.16.10.11',
                      description='Controller Floor 2', site='NY2',
                      site_description='Secondary', contract='NY-002'),
        base._replace(hostname='fw', ip='10.10.20.20',
                      description='Firewall'),
    )
    for row in rows:
        cis_sheet.append(row)


def set_appliances_sheet(workbook):
    appliances_sheet = workbook.create_sheet(APPLIANCES_SHEET)
    Row = namedtuple('ApplianceRow', (
        'ci_hostname', 'serial_number', 'manufacture', 'model', 'virtual',
    ))
    rows = (
        Row._fields,
        Row('wlc1', 'FOX123', 'Cisco', '3560', 'x'),
        Row('wlc1', 'FOX124', 'Cisco', '3560', 'x'),
        Row('wlc2', 'FOX125', 'Cisco', '3560', 'x'),
        Row('router_sp', 'TYF987', 'Cisco', '2960', 'x'),
        Row('router_bh', 'TYF654', 'Cisco', '2960', 'x'),
        Row('fw', '687F', 'F5', 'BIG-IP', ''),
    )
    for row in rows:
        appliances_sheet.append(row)
