import logging

from collections import namedtuple
from openpyxl import load_workbook
from django.db import IntegrityError, transaction
from typing import Set

from .models import Client, Place, CI, Appliance, Contract, Manufacturer
from .cis_mapping import HOSTNAME, IP, DESCRIPTION, \
    DEPLOYED, BUSINESS_IMPACT, PLACE, PLACE_DESCRIPTION, CONTRACT, \
    CONTRACT_BEGIN, CONTRACT_END, CONTRACT_DESCRIPTION, CREDENTIAL_USERNAME, \
    CREDENTIAL_PASSWORD, CREDENTIAL_ENABLE_PASSWORD, CREDENTIAL_INSTRUCTIONS, \
    CIS_SHEET, APPLIANCES_SHEET, APPLIANCE_HOSTNAME, APPLIANCE_SERIAL_NUMBER, \
    APPLIANCE_MANUFACTURER, APPLIANCE_MODEL, APPLIANCE_VIRTUAL


logger = logging.getLogger(__name__)


class CILoader:
    def __init__(self, file, client: Client):
        self._workbook = load_workbook(file, read_only=True, data_only=True)
        self.client = client
        self.places = {}
        self.contracts = {}
        self.manufacturers = {}
        self.cis = []
        self.errors = []

    def save(self):
        cis_sheet = self._workbook[CIS_SHEET]
        logger.info(f'The method save() of the class {self.__class__.__name__} was called.')

        Error = namedtuple('Error', ['exc', 'row'])
        for row in cis_sheet.iter_rows(min_row=2, values_only=True):
            try:
                with transaction.atomic():
                    ci = self._create_ci(row)
                    ci.appliances.set(self._get_ci_appliances(row[HOSTNAME]))
                self.cis.append(ci)
                logger.info(f'{ci} was added to self.cis')
            except IntegrityError as e:
                self.errors.append(Error(e, row))
                logger.error(f'{e} spreadsheet row: {row} was added to self.errors')
        return self

    def _create_ci(self, row: tuple) -> CI:
        return CI.objects.create(
            client=self.client,
            hostname=row[HOSTNAME],
            ip=row[IP],
            description=row[DESCRIPTION],
            deployed=bool(row[DEPLOYED]),
            business_impact=self._get_business_impact(row[BUSINESS_IMPACT]),
            place=self._get_place(row[PLACE], row[PLACE_DESCRIPTION]),
            contract=self._get_contract(row),
            username=row[CREDENTIAL_USERNAME],
            password=row[CREDENTIAL_PASSWORD],
            enable_password=row[CREDENTIAL_ENABLE_PASSWORD],
            instructions=row[CREDENTIAL_INSTRUCTIONS],
        )

    def _get_ci_appliances(self, hostname: str) -> Set[Appliance]:
        appliances = set()
        appliances_sheet = self._workbook[APPLIANCES_SHEET]
        for appl_row in appliances_sheet.iter_rows(min_row=2, values_only=True):
            if appl_row[APPLIANCE_HOSTNAME] == hostname:
                appliances.add(self._get_appliance(appl_row))
        return appliances

    def _get_place(self, name: str, description: str) -> Place:
        if name in self.places:
            return self.places[name]
        else:
            self.places[name] = Place.objects.get_or_create(
                name=name,
                description=description,
                client=self.client
            )[0]
            return self.places[name]

    def _get_contract(self, row: tuple) -> Contract:
        contract_name = row[CONTRACT]
        if contract_name in self.contracts:
            return self.contracts[contract_name]
        else:
            self.contracts[contract_name] = Contract.objects.get_or_create(
                description=row[CONTRACT_DESCRIPTION],
                name=contract_name,
                begin=row[CONTRACT_BEGIN],
                end=row[CONTRACT_END],
            )[0]
            return self.contracts[contract_name]

    def _get_appliance(self, row) -> Appliance:
        appliance = Appliance.objects.get_or_create(
            client=self.client,
            serial_number=row[APPLIANCE_SERIAL_NUMBER],
            manufacturer=self._get_manufacturer(row[APPLIANCE_MANUFACTURER]),
            model=row[APPLIANCE_MODEL],
            virtual=bool(str(row[APPLIANCE_VIRTUAL]).strip())
        )[0]
        return appliance

    def _get_manufacturer(self, name: str) -> Manufacturer:
        if name in self.manufacturers:
            return self.manufacturers[name]
        else:
            self.manufacturers[name] = Manufacturer.objects.get_or_create(
                name=name,
            )[0]
            return self.manufacturers[name]

    @staticmethod
    def _get_business_impact(business_impact: str) -> int:
        model_choices = dict(CI.IMPACT_OPTIONS).items()
        options = {value: key for key, value in model_choices}
        return options.get(business_impact.lower())
